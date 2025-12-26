# app/utils/export.py
import os
import csv
import logging
from typing import List, Dict, Any, Optional, Union, IO
from pathlib import Path
from datetime import datetime
from io import StringIO, BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd

from app.core.config import settings

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exportateur Excel avancé"""
    
    def __init__(self):
        self.output_dir = Path(settings.MEDIA_ROOT) / "exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "Données",
        include_header: bool = True,
        auto_width: bool = True,
        style: str = "default"
    ) -> str:
        """
        Exporte des données vers un fichier Excel
        
        Args:
            data: Liste de dictionnaires
            filename: Nom du fichier (sans extension)
            sheet_name: Nom de la feuille
            include_header: Inclure les en-têtes
            auto_width: Ajuster automatiquement la largeur des colonnes
            style: Style d'export (default, minimal, corporate)
        
        Returns:
            Chemin du fichier généré
        """
        try:
            # Créer un nouveau classeur
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            if not data:
                # Feuille vide avec message
                ws["A1"] = "Aucune donnée à exporter"
                ws["A1"].font = Font(italic=True, color="808080")
            else:
                # Écrire les en-têtes
                if include_header and data:
                    headers = list(data[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        self._apply_header_style(cell, style)
                
                # Écrire les données
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, key in enumerate(headers, 1):
                        value = row_data.get(key)
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        self._apply_cell_style(cell, style, row_idx)
                
                # Ajuster la largeur des colonnes
                if auto_width:
                    self._auto_adjust_column_width(ws)
                
                # Appliquer des styles selon le type
                self._apply_workbook_styles(wb, style)
            
            # Sauvegarder le fichier
            filepath = self.output_dir / f"{filename}.xlsx"
            wb.save(str(filepath))
            
            logger.info(f"Fichier Excel généré: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur export Excel: {str(e)}")
            raise
    
    def export_to_excel_bytes(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Données"
    ) -> BytesIO:
        """Exporte vers un BytesIO (pour téléchargement direct)"""
        output = BytesIO()
        
        if data:
            # Utiliser pandas pour une conversion simple
            df = pd.DataFrame(data)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # Classeur vide
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws["A1"] = "Aucune donnée"
            wb.save(output)
        
        output.seek(0)
        return output
    
    def _apply_header_style(self, cell, style: str):
        """Applique le style aux en-têtes"""
        if style == "corporate":
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="366092",  # Bleu foncé
                end_color="366092",
                fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif style == "minimal":
            cell.font = Font(bold=True)
        else:  # default
            cell.font = Font(bold=True, color="000000")
            cell.fill = PatternFill(
                start_color="F2F2F2",
                end_color="F2F2F2",
                fill_type="solid"
            )
    
    def _apply_cell_style(self, cell, style: str, row_idx: int):
        """Applique le style aux cellules de données"""
        if style == "corporate":
            # Lignes alternées
            if row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F8F9FA",
                    end_color="F8F9FA",
                    fill_type="solid"
                )
            
            # Bordures fines
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
            
        elif style == "minimal":
            # Style minimal, juste des bordures bas
            cell.border = Border(bottom=Side(style='thin'))
    
    def _auto_adjust_column_width(self, ws):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Limite à 50 caractères
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_workbook_styles(self, wb, style: str):
        """Applique des styles au niveau du classeur"""
        if style == "corporate":
            # Définir les propriétés du document
            wb.properties.title = "Export de données"
            wb.properties.subject = "Export automatique"
            wb.properties.creator = settings.PROJECT_NAME
            wb.properties.keywords = "export, données"


class CSVExporter:
    """Exportateur CSV"""
    
    def __init__(self):
        self.output_dir = Path(settings.MEDIA_ROOT) / "exports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        delimiter: str = ";",
        encoding: str = "utf-8-sig"
    ) -> str:
        """
        Exporte des données vers un fichier CSV
        
        Args:
            data: Liste de dictionnaires
            filename: Nom du fichier (sans extension)
            delimiter: Séparateur de champ
            encoding: Encodage du fichier
        
        Returns:
            Chemin du fichier généré
        """
        try:
            if not data:
                raise ValueError("Aucune donnée à exporter")
            
            filepath = self.output_dir / f"{filename}.csv"
            
            # Déterminer les en-têtes
            headers = list(data[0].keys())
            
            with open(filepath, 'w', newline='', encoding=encoding) as csvfile:
                writer = csv.DictWriter(
                    csvfile,
                    fieldnames=headers,
                    delimiter=delimiter,
                    quotechar='"',
                    quoting=csv.QUOTE_MINIMAL
                )
                writer.writeheader()
                writer.writerows(data)
            
            logger.info(f"Fichier CSV généré: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Erreur export CSV: {str(e)}")
            raise
    
    def export_to_csv_string(
        self,
        data: List[Dict[str, Any]],
        delimiter: str = ";"
    ) -> str:
        """Exporte vers une chaîne CSV"""
        output = StringIO()
        
        if data:
            headers = list(data[0].keys())
            writer = csv.DictWriter(
                output,
                fieldnames=headers,
                delimiter=delimiter
            )
            writer.writeheader()
            writer.writerows(data)
        
        return output.getvalue()
    
    def export_to_csv_bytes(
        self,
        data: List[Dict[str, Any]],
        delimiter: str = ";",
        encoding: str = "utf-8"
    ) -> BytesIO:
        """Exporte vers BytesIO"""
        output = BytesIO()
        
        if data:
            headers = list(data[0].keys())
            # Écrire BOM pour Excel
            if encoding == "utf-8-sig":
                output.write(b'\xef\xbb\xbf')
            
            writer = csv.DictWriter(
                output,
                fieldnames=headers,
                delimiter=delimiter
            )
            writer.writeheader()
            for row in data:
                writer.writerow(row)
        
        output.seek(0)
        return output


def export_to_excel(
    data: List[Dict[str, Any]],
    filename: str,
    sheet_name: str = "Données",
    format: str = "xlsx",
    **kwargs
) -> Union[str, BytesIO]:
    """
    Fonction utilitaire pour exporter vers Excel ou CSV
    
    Args:
        data: Données à exporter
        filename: Nom du fichier
        sheet_name: Nom de la feuille (Excel uniquement)
        format: Format d'export (xlsx, csv)
        **kwargs: Arguments supplémentaires
    
    Returns:
        Chemin du fichier ou BytesIO
    """
    if format.lower() == "csv":
        exporter = CSVExporter()
        return exporter.export_to_csv(data, filename, **kwargs)
    else:
        exporter = ExcelExporter()
        return exporter.export_to_excel(data, filename, sheet_name, **kwargs)


def generate_export_filename(
    prefix: str,
    entity_type: str,
    extension: str = "xlsx"
) -> str:
    """
    Génère un nom de fichier d'export standardisé
    
    Args:
        prefix: Préfixe (ex: "ventes", "clients")
        entity_type: Type d'entité
        extension: Extension du fichier
    
    Returns:
        Nom de fichier généré
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{entity_type}_{timestamp}.{extension}"


def cleanup_old_exports(days_old: int = 7):
    """
    Nettoie les anciens fichiers d'export
    
    Args:
        days_old: Âge maximum des fichiers (jours)
    """
    try:
        export_dir = Path(settings.MEDIA_ROOT) / "exports"
        cutoff_time = datetime.now().timestamp() - (days_old * 24 * 60 * 60)
        
        deleted_count = 0
        for filepath in export_dir.glob("*.xlsx"):
            if filepath.stat().st_mtime < cutoff_time:
                filepath.unlink()
                deleted_count += 1
        
        for filepath in export_dir.glob("*.csv"):
            if filepath.stat().st_mtime < cutoff_time:
                filepath.unlink()
                deleted_count += 1
        
        logger.info(f"{deleted_count} anciens fichiers d'export supprimés")
        
    except Exception as e:
        logger.error(f"Erreur nettoyage exports: {str(e)}")