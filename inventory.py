# app/api/routes/inventory.py
from fastapi import APIRouter, Depends, HTTPException, status, Query, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from typing import List, Optional, Dict, Any
from uuid import UUID
from datetime import datetime, date, timedelta
import logging
from pathlib import Path
from app.db.session import get_db
from app.models.inventory import PhysicalInventory, InventoryItem, InventorySchedule
from app.models.product import Product
from app.models.tenant import Tenant
from app.models.user import User
from app.schemas.inventory import (
    InventoryCreate, InventoryInDB, InventoryUpdate,
    InventoryItemCreate, InventoryItemInDB, InventoryItemUpdate,
    InventoryReport, ScheduleCreate
)
from app.api.deps import get_current_tenant, get_current_user
from app.core.security import require_permission
from app.services.inventory import InventoryService

router = APIRouter(prefix="/inventory", tags=["Inventory"])
logger = logging.getLogger(__name__)

@router.post("/", response_model=InventoryInDB)
@require_permission("inventory_manage")
def create_inventory(
    inventory_data: InventoryCreate,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Crée un nouvel inventaire physique
    """
    try:
        # Générer le numéro d'inventaire
        inventory_number = f"INV-{datetime.now().strftime('%Y%m%d')}-{UUID().hex[:8].upper()}"
        
        # Créer l'inventaire
        inventory = PhysicalInventory(
            tenant_id=current_tenant.id,
            inventory_number=inventory_number,
            inventory_type=inventory_data.inventory_type.value,
            description=inventory_data.description,
            planned_date=inventory_data.planned_date,
            tags=inventory_data.tags,
            created_by=current_user.id,
            status="draft"
        )
        
        db.add(inventory)
        db.flush()  # Pour obtenir l'ID
        
        # Ajouter les items selon le type d'inventaire
        if inventory_data.inventory_type == "partial" and inventory_data.product_ids:
            # Inventaire partiel - seulement les produits spécifiés
            products = db.query(Product).filter(
                Product.tenant_id == current_tenant.id,
                Product.id.in_(inventory_data.product_ids)
            ).all()
        else:
            # Inventaire complet - tous les produits actifs
            products = db.query(Product).filter(
                Product.tenant_id == current_tenant.id,
                Product.is_active == True
            ).all()
        
        # Créer les items d'inventaire
        for product in products:
            item = InventoryItem(
                tenant_id=current_tenant.id,
                inventory_id=inventory.id,
                product_id=product.id,
                expected_quantity=product.quantity,
                expected_value=product.quantity * product.purchase_price,
                status="pending"
            )
            db.add(item)
        
        db.commit()
        db.refresh(inventory)
        
        logger.info(f"Inventaire créé: {inventory_number} par {current_user.nom_complet}")
        
        return inventory
        
    except Exception as e:
        db.rollback()
        logger.error(f"Erreur lors de la création de l'inventaire: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la création de l'inventaire"
        )

@router.get("/", response_model=List[InventoryInDB])
@require_permission("inventory_view")
def list_inventories(
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status: Optional[str] = None,
    inventory_type: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
):
    """
    Liste les inventaires avec filtres
    """
    query = db.query(PhysicalInventory).filter(
        PhysicalInventory.tenant_id == current_tenant.id
    )
    
    # Appliquer les filtres
    if status:
        query = query.filter(PhysicalInventory.status == status)
    
    if inventory_type:
        query = query.filter(PhysicalInventory.inventory_type == inventory_type)
    
    if start_date:
        query = query.filter(PhysicalInventory.start_date >= start_date)
    
    if end_date:
        query = query.filter(PhysicalInventory.start_date <= end_date)
    
    # Trier par date de création décroissante
    inventories = query.order_by(
        PhysicalInventory.created_at.desc()
    ).offset(skip).limit(limit).all()
    
    return inventories

@router.get("/{inventory_id}", response_model=InventoryReport)
@require_permission("inventory_view")
def get_inventory(
    inventory_id: UUID,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Récupère un inventaire avec ses items
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    # Récupérer les items avec les informations des produits
    items = db.query(
        InventoryItem,
        Product.name.label('product_name'),
        Product.code.label('product_code')
    ).join(
        Product, InventoryItem.product_id == Product.id
    ).filter(
        InventoryItem.inventory_id == inventory_id,
        InventoryItem.tenant_id == current_tenant.id
    ).all()
    
    # Formater les items
    inventory_items = []
    for item, product_name, product_code in items:
        item_dict = InventoryItemInDB.from_orm(item)
        item_dict.product_name = product_name
        item_dict.product_code = product_code
        inventory_items.append(item_dict)
    
    # Calculer le résumé
    summary = {
        "total_items": inventory.total_items,
        "items_counted": inventory.items_counted,
        "items_missing": inventory.items_missing,
        "items_excess": inventory.items_excess,
        "completion_rate": (inventory.items_counted / inventory.total_items * 100) if inventory.total_items > 0 else 0,
        "system_value": inventory.system_value,
        "counted_value": inventory.counted_value,
        "variance_value": inventory.variance_value,
        "variance_percentage": inventory.variance_percentage
    }
    
    # Générer des recommandations
    recommendations = []
    if inventory.variance_percentage > 5:
        recommendations.append("Écart significatif détecté. Vérifier les procédures de stockage.")
    if inventory.items_missing > 0:
        recommendations.append(f"{inventory.items_missing} items manquants. Investigation requise.")
    if inventory.items_counted < inventory.total_items:
        recommendations.append(f"Inventaire incomplet: {inventory.total_items - inventory.items_counted} items restants.")
    
    return InventoryReport(
        inventory=inventory,
        items=inventory_items,
        summary=summary,
        recommendations=recommendations
    )

@router.post("/{inventory_id}/items")
@require_permission("inventory_manage")
def add_inventory_item(
    inventory_id: UUID,
    item_data: InventoryItemCreate,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Ajoute un item à un inventaire
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    if inventory.status not in ["draft", "in_progress"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Inventaire non modifiable"
        )
    
    product = db.query(Product).filter(
        Product.id == item_data.product_id,
        Product.tenant_id == current_tenant.id
    ).first()
    
    if not product:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Produit non trouvé"
        )
    
    try:
        # Vérifier si l'item existe déjà
        existing_item = db.query(InventoryItem).filter(
            InventoryItem.inventory_id == inventory_id,
            InventoryItem.product_id == item_data.product_id,
            InventoryItem.tenant_id == current_tenant.id
        ).first()
        
        if existing_item:
            # Mettre à jour l'item existant
            existing_item.counted_quantity = item_data.counted_quantity
            existing_item.counted_at = datetime.utcnow()
            existing_item.batch_number = item_data.batch_number
            existing_item.expiry_date = item_data.expiry_date
            existing_item.location = item_data.location
            existing_item.notes = item_data.notes
            existing_item.status = "counted"
            
            # Calculer les variances
            existing_item.calculate_variance()
        else:
            # Créer un nouvel item
            item = InventoryItem(
                tenant_id=current_tenant.id,
                inventory_id=inventory_id,
                product_id=item_data.product_id,
                expected_quantity=product.quantity,
                counted_quantity=item_data.counted_quantity,
                batch_number=item_data.batch_number,
                expiry_date=item_data.expiry_date,
                location=item_data.location,
                notes=item_data.notes,
                counted_at=datetime.utcnow(),
                status="counted"
            )
            item.calculate_variance()
            db.add(item)
        
        # Mettre à jour les statistiques de l'inventaire
        inventory.calculate_variance()
        
        db.commit()
        
        logger.info(f"Item ajouté à l'inventaire {inventory.inventory_number}")
        
        return {"message": "Item ajouté avec succès"}
        
    except Exception as e:
        db.rollback()
        logger.error(f"Erreur lors de l'ajout de l'item: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de l'ajout de l'item"
        )

@router.post("/{inventory_id}/start")
@require_permission("inventory_manage")
def start_inventory(
    inventory_id: UUID,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Démarre un inventaire
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    if inventory.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="L'inventaire a déjà été démarré"
        )
    
    try:
        inventory.status = "in_progress"
        inventory.start_date = datetime.utcnow()
        db.commit()
        
        logger.info(f"Inventaire démarré: {inventory.inventory_number}")
        
        return {"message": "Inventaire démarré avec succès"}
        
    except Exception as e:
        db.rollback()
        logger.error(f"Erreur lors du démarrage de l'inventaire: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors du démarrage de l'inventaire"
        )

@router.post("/{inventory_id}/complete")
@require_permission("inventory_manage")
def complete_inventory(
    inventory_id: UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Termine un inventaire et ajuste les stocks
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    if inventory.status not in ["in_progress", "counting"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="L'inventaire n'est pas en cours"
        )
    
    try:
        # Marquer l'inventaire comme terminé
        inventory.status = "completed"
        inventory.end_date = datetime.utcnow()
        inventory.validated_by = current_user.id
        
        # Ajuster les stocks pour les items avec variances
        for item in inventory.items:
            if item.variance != 0 and item.product:
                # Créer un mouvement de stock pour l'ajustement
                from app.models.stock_movement import StockMovement
                
                movement = StockMovement(
                    tenant_id=current_tenant.id,
                    product_id=item.product.id,
                    quantity_before=item.expected_quantity,
                    quantity_after=item.counted_quantity,
                    quantity_change=item.variance,
                    movement_type="inventory_adjustment",
                    reason=f"Ajustement d'inventaire {inventory.inventory_number}",
                    reference_number=inventory.inventory_number,
                    created_by=current_user.id
                )
                
                # Mettre à jour le stock du produit
                item.product.quantity = item.counted_quantity
                
                db.add(movement)
        
        db.commit()
        
        # Lancer la génération du rapport en arrière-plan
        background_tasks.add_task(
            generate_inventory_report,
            inventory_id=inventory_id,
            tenant_id=current_tenant.id
        )
        
        logger.info(f"Inventaire terminé: {inventory.inventory_number}")
        
        return {
            "message": "Inventaire terminé avec succès",
            "variance_value": inventory.variance_value,
            "variance_percentage": inventory.variance_percentage
        }
        
    except Exception as e:
        db.rollback()
        logger.error(f"Erreur lors de la finalisation de l'inventaire: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la finalisation de l'inventaire"
        )

@router.get("/{inventory_id}/export")
@require_permission("inventory_view")
def export_inventory(
    inventory_id: UUID,
    export_format: str = Query("excel", pattern="^(excel|pdf|csv)$"),
    background_tasks: BackgroundTasks = None,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Exporte un inventaire dans différents formats
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    if background_tasks:
        # Lancer l'export en arrière-plan
        from app.services.export import ExportService
        export_service = ExportService(current_tenant)
        
        background_tasks.add_task(
            export_service.export_inventory,
            inventory_id=inventory_id,
            export_format=export_format,
            user_id=current_user.id
        )
        
        return {
            "message": "Export démarré en arrière-plan",
            "format": export_format,
            "inventory_number": inventory.inventory_number
        }
    
    # Retour direct pour petits exports
    return {"message": "Export synchrone non implémenté"}

@router.post("/schedules")
@require_permission("inventory_manage")
def create_inventory_schedule(
    schedule_data: ScheduleCreate,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Crée un planning d'inventaire récurrent
    """
    try:
        # Calculer la prochaine date
        today = date.today()
        next_schedule = today
        
        if schedule_data.schedule_type == "daily":
            next_schedule = today + timedelta(days=schedule_data.frequency)
        elif schedule_data.schedule_type == "weekly":
            days_ahead = schedule_data.day_of_week - today.weekday()
            if days_ahead <= 0:
                days_ahead += 7
            next_schedule = today + timedelta(days=days_ahead)
        elif schedule_data.schedule_type == "monthly":
            year = today.year
            month = today.month
            
            if today.day >= schedule_data.day_of_month:
                month += 1
                if month > 12:
                    month = 1
                    year += 1
            
            next_schedule = date(year, month, min(schedule_data.day_of_month, 28))
        elif schedule_data.schedule_type == "yearly":
            year = today.year
            if today.month > schedule_data.month_of_year or (
                today.month == schedule_data.month_of_year and today.day >= 1
            ):
                year += 1
            next_schedule = date(year, schedule_data.month_of_year, 1)
        
        schedule = InventorySchedule(
            tenant_id=current_tenant.id,
            schedule_type=schedule_data.schedule_type.value,
            frequency=schedule_data.frequency,
            day_of_week=schedule_data.day_of_week,
            day_of_month=schedule_data.day_of_month,
            month_of_year=schedule_data.month_of_year,
            cycle_count=schedule_data.cycle_count or 0,
            description=schedule_data.description,
            next_schedule=next_schedule
        )
        
        db.add(schedule)
        db.commit()
        
        logger.info(f"Planning d'inventaire créé par {current_user.nom_complet}")
        
        return {"message": "Planning créé avec succès", "next_schedule": next_schedule}
        
    except Exception as e:
        db.rollback()
        logger.error(f"Erreur lors de la création du planning: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Erreur lors de la création du planning"
        )

# Ajoutez cette fonction dans app/api/routes/inventory.py, après les autres routes

async def generate_inventory_report(
    inventory_id: UUID,
    tenant_id: UUID,
    db: Session = None
):
    """
    Génère un rapport d'inventaire en arrière-plan
    """
    try:
        if db is None:
            from app.db.session import get_db
            db = next(get_db())
        
        # Récupérer l'inventaire avec ses items
        inventory = db.query(PhysicalInventory).filter(
            PhysicalInventory.id == inventory_id,
            PhysicalInventory.tenant_id == tenant_id
        ).first()
        
        if not inventory:
            logger.error(f"Inventaire {inventory_id} non trouvé")
            return
        
        # Récupérer les items avec les produits
        items_with_products = db.query(
            InventoryItem,
            Product
        ).join(
            Product, InventoryItem.product_id == Product.id
        ).filter(
            InventoryItem.inventory_id == inventory_id,
            InventoryItem.tenant_id == tenant_id
        ).all()
        
        # Préparer les données du rapport
        report_data = {
            "inventory_id": str(inventory.id),
            "inventory_number": inventory.inventory_number,
            "inventory_type": inventory.inventory_type,
            "status": inventory.status,
            "created_at": inventory.created_at.isoformat() if inventory.created_at else None,
            "start_date": inventory.start_date.isoformat() if inventory.start_date else None,
            "end_date": inventory.end_date.isoformat() if inventory.end_date else None,
            "total_items": inventory.total_items,
            "items_counted": inventory.items_counted,
            "items_missing": inventory.items_missing,
            "items_excess": inventory.items_excess,
            "completion_rate": inventory.items_counted / inventory.total_items * 100 if inventory.total_items > 0 else 0,
            "system_value": float(inventory.system_value) if inventory.system_value else 0,
            "counted_value": float(inventory.counted_value) if inventory.counted_value else 0,
            "variance_value": float(inventory.variance_value) if inventory.variance_value else 0,
            "variance_percentage": float(inventory.variance_percentage) if inventory.variance_percentage else 0,
            "items": []
        }
        
        # Ajouter les items
        for item, product in items_with_products:
            item_data = {
                "product_id": str(product.id),
                "product_code": product.code,
                "product_name": product.name,
                "expected_quantity": float(item.expected_quantity) if item.expected_quantity else 0,
                "counted_quantity": float(item.counted_quantity) if item.counted_quantity else 0,
                "variance": float(item.variance) if item.variance else 0,
                "variance_percentage": float(item.variance_percentage) if item.variance_percentage else 0,
                "batch_number": item.batch_number,
                "expiry_date": item.expiry_date.isoformat() if item.expiry_date else None,
                "location": item.location,
                "notes": item.notes,
                "status": item.status,
                "counted_at": item.counted_at.isoformat() if item.counted_at else None
            }
            report_data["items"].append(item_data)
        
        # Générer le rapport (exemple avec création de fichier)
        import json
        from pathlib import Path
        from datetime import datetime
        
        # Créer le dossier de rapports
        reports_dir = Path("reports/inventory")
        reports_dir.mkdir(parents=True, exist_ok=True)
        
        # Nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_report_{inventory.inventory_number}_{timestamp}.json"
        filepath = reports_dir / filename
        
        # Sauvegarder le rapport JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        # Option: Générer un PDF ou Excel
        try:
            # Générer un rapport Excel
            await generate_inventory_excel_report(report_data, filepath.with_suffix('.xlsx'))
        except Exception as e:
            logger.warning(f"Échec génération Excel: {str(e)}")
        
        logger.info(f"Rapport d'inventaire généré: {filepath}")
        
        # Mettre à jour l'inventaire avec le chemin du rapport
        inventory.report_path = str(filepath)
        db.commit()
        
        return report_data
        
    except Exception as e:
        logger.error(f"Erreur génération rapport inventaire {inventory_id}: {str(e)}")
        raise


async def generate_inventory_excel_report(report_data: dict, output_path: Path):
    """
    Génère un rapport Excel pour l'inventaire
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport Inventaire"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # En-tête du rapport
        ws['A1'] = "RAPPORT D'INVENTAIRE"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:H1')
        
        # Informations de l'inventaire
        ws['A3'] = "Numéro d'inventaire:"
        ws['B3'] = report_data.get('inventory_number', 'N/A')
        
        ws['A4'] = "Type:"
        ws['B4'] = report_data.get('inventory_type', 'N/A')
        
        ws['A5'] = "Statut:"
        ws['B5'] = report_data.get('status', 'N/A')
        
        ws['A6'] = "Date de début:"
        ws['B6'] = report_data.get('start_date', 'N/A')
        
        ws['A7'] = "Date de fin:"
        ws['B7'] = report_data.get('end_date', 'N/A')
        
        # Résumé
        ws['A9'] = "RÉSUMÉ"
        ws['A9'].font = Font(bold=True, size=14)
        ws.merge_cells('A9:H9')
        
        summary_rows = [
            ("Total items", report_data.get('total_items', 0)),
            ("Items comptés", report_data.get('items_counted', 0)),
            ("Items manquants", report_data.get('items_missing', 0)),
            ("Items en excès", report_data.get('items_excess', 0)),
            ("Taux de complétion", f"{report_data.get('completion_rate', 0):.1f}%"),
            ("Valeur système", f"{report_data.get('system_value', 0):.2f}"),
            ("Valeur comptée", f"{report_data.get('counted_value', 0):.2f}"),
            ("Écart valeur", f"{report_data.get('variance_value', 0):.2f}"),
            ("Écart %", f"{report_data.get('variance_percentage', 0):.2f}%")
        ]
        
        for i, (label, value) in enumerate(summary_rows, start=10):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
        
        # Tableau des items
        ws['A20'] = "DÉTAIL DES ITEMS"
        ws['A20'].font = Font(bold=True, size=14)
        ws.merge_cells('A20:K20')
        
        # En-têtes du tableau
        headers = [
            "Code Produit",
            "Nom Produit",
            "Quantité Attendue",
            "Quantité Comptée",
            "Écart",
            "Écart %",
            "N° Lot",
            "Date Péremption",
            "Localisation",
            "Statut",
            "Notes"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=22, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Données des items
        items = report_data.get('items', [])
        for row_idx, item in enumerate(items, start=23):
            ws.cell(row=row_idx, column=1, value=item.get('product_code', ''))
            ws.cell(row=row_idx, column=2, value=item.get('product_name', ''))
            ws.cell(row=row_idx, column=3, value=item.get('expected_quantity', 0))
            ws.cell(row=row_idx, column=4, value=item.get('counted_quantity', 0))
            
            variance = item.get('variance', 0)
            variance_cell = ws.cell(row=row_idx, column=5, value=variance)
            if variance < 0:
                variance_cell.font = Font(color="FF0000")  # Rouge pour déficit
            elif variance > 0:
                variance_cell.font = Font(color="00B050")  # Vert pour excès
            
            ws.cell(row=row_idx, column=6, value=f"{item.get('variance_percentage', 0):.2f}%")
            ws.cell(row=row_idx, column=7, value=item.get('batch_number', ''))
            ws.cell(row=row_idx, column=8, value=item.get('expiry_date', ''))
            ws.cell(row=row_idx, column=9, value=item.get('location', ''))
            ws.cell(row=row_idx, column=10, value=item.get('status', ''))
            ws.cell(row=row_idx, column=11, value=item.get('notes', ''))
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sauvegarder le fichier
        wb.save(output_path)
        logger.info(f"Rapport Excel généré: {output_path}")
        
    except ImportError:
        logger.warning("openpyxl non installé. Impossible de générer le rapport Excel.")
    except Exception as e:
        logger.error(f"Erreur génération Excel: {str(e)}")


# Ajoutez également cette route pour télécharger le rapport
@router.get("/{inventory_id}/report")
@require_permission("inventory_view")
async def download_inventory_report(
    inventory_id: UUID,
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user)
):
    """
    Télécharge le rapport d'inventaire
    """
    inventory = db.query(PhysicalInventory).filter(
        PhysicalInventory.id == inventory_id,
        PhysicalInventory.tenant_id == current_tenant.id
    ).first()
    
    if not inventory:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Inventaire non trouvé"
        )
    
    if not inventory.report_path:
        # Générer le rapport si nécessaire
        report_data = await generate_inventory_report(inventory_id, current_tenant.id, db)
        
        from fastapi.responses import JSONResponse
        return JSONResponse(
            content=report_data,
            headers={"Content-Disposition": f"attachment; filename=inventory_report_{inventory.inventory_number}.json"}
        )
    
    # Retourner le fichier existant
    import os
    from fastapi.responses import FileResponse
    
    if os.path.exists(inventory.report_path):
        filename = f"inventory_report_{inventory.inventory_number}.json"
        return FileResponse(
            path=inventory.report_path,
            filename=filename,
            media_type="application/json"
        )
    
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail="Rapport non disponible"
    )


# Ajoutez aussi cette route pour les statistiques d'inventaire
@router.get("/stats/summary")
@require_permission("inventory_view")
def get_inventory_stats(
    db: Session = Depends(get_db),
    current_tenant: Tenant = Depends(get_current_tenant),
    current_user: User = Depends(get_current_user),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None)
):
    """
    Statistiques des inventaires
    """
    query = db.query(PhysicalInventory).filter(
        PhysicalInventory.tenant_id == current_tenant.id,
        PhysicalInventory.status == "completed"
    )
    
    if start_date:
        query = query.filter(PhysicalInventory.end_date >= start_date)
    
    if end_date:
        query = query.filter(PhysicalInventory.end_date <= end_date)
    
    inventories = query.all()
    
    if not inventories:
        return {
            "total_inventories": 0,
            "total_items": 0,
            "average_variance": 0,
            "total_variance_value": 0
        }
    
    stats = {
        "total_inventories": len(inventories),
        "total_items": sum(inv.total_items for inv in inventories),
        "total_variance_value": float(sum(inv.variance_value for inv in inventories if inv.variance_value)),
        "average_variance": float(sum(inv.variance_percentage for inv in inventories if inv.variance_percentage) / len(inventories)),
        "inventories_by_type": {},
        "recent_inventories": []
    }
    
    # Distribution par type
    for inv in inventories:
        inv_type = inv.inventory_type or "unknown"
        stats["inventories_by_type"][inv_type] = stats["inventories_by_type"].get(inv_type, 0) + 1
    
    # 5 derniers inventaires
    recent_inventories = sorted(inventories, key=lambda x: x.end_date or x.created_at, reverse=True)[:5]
    stats["recent_inventories"] = [
        {
            "id": str(inv.id),
            "number": inv.inventory_number,
            "type": inv.inventory_type,
            "end_date": inv.end_date.isoformat() if inv.end_date else None,
            "variance_percentage": float(inv.variance_percentage) if inv.variance_percentage else 0,
            "total_items": inv.total_items
        }
        for inv in recent_inventories
    ]
    
    return stats
